import openpyxl

workbook_path = "/Users/maxmaton/Library/Mobile Documents/com~apple~CloudDocs/Documents/Skills & Hobbies/Programming/Python/Open PyXL/OpenPyXL practice workbook.xlsx"

wb = openpyxl.load_workbook(workbook_path)

# Creating a function that allows you to copy and paste a range to a selected location

def copy_paste_cell_range(source_sheet_input, source_range_input, dest_range_start, dest_file_input = None, dest_sheet_input = None):

    source_workbook = wb

    source_sheet = source_workbook[source_sheet_input]

    source_col_start, source_row_start, source_col_end, source_row_end = openpyxl.utils.range_boundaries(source_range_input)


    if dest_file_input is None:
        dest_file_input = workbook_path

    dest_workbook = openpyxl.load_workbook(dest_file_input)

    if dest_sheet_input is None:
        dest_sheet_input = source_sheet_input


    dest_col_start = dest_range_start.rstrip('0123456789')
    dest_row_start = int(dest_range_start[len(dest_col_start):])

    dest_col_start = openpyxl.utils.column_index_from_string(dest_col_start)

    dest_sheet = dest_workbook[dest_sheet_input]

    dest_row = dest_row_start
    for row in range(source_row_start, source_row_end + 1):

        dest_col = dest_col_start

        for col in range(source_col_start, source_col_end + 1):

            source_cell = source_sheet.cell(row, col)
            dest_cell = dest_sheet.cell(dest_row, dest_col)

            dest_cell.value = source_cell.value

            dest_col += 1

        dest_row +=1

    dest_workbook.save(dest_file_input)

# Copying and pasting in the same worksheet in the same workbook:

copy_paste_cell_range(source_sheet_input = "Data", source_range_input = "A1:G10", dest_range_start = "A12")