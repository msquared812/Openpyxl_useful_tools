import openpyxl
import os

# setting the working directory

os.chdir("/Users/maxmaton/Library/Mobile Documents/com~apple~CloudDocs/Documents/Skills & Hobbies/Programming/Python/Open PyXL/")

# creating a blank workbook

wb_new = openpyxl.Workbook()

wb_new.save('New workbook.xlsx')

# copying workbooks n number of times, where the default value for n is 1:

def copy_workbook(filename, n = 1):

    if type(filename) != str:
        return "Please input a string for the filename"

    filename_and_extension = filename.split('.')

    workbook_to_copy = openpyxl.load_workbook(filename)

    for i in range(1, n + 1):

        workbook_to_copy.save(filename_and_extension[0] + " copy " + str(i) + "." + filename_and_extension[1])


copy_workbook("New workbook.xlsx", 2)

# Creating new sheets

wb = openpyxl.load_workbook("OpenPyXL practice workbook.xlsx")

print(wb.sheetnames)

wb.create_sheet("Sheet 2")

# Creating new sheets with names from a list of strings

list_strings = ["Spain", "Mexico", "France"]

for name in list_strings:
    wb.create_sheet(name)

wb.save("OpenPyXL practice workbook.xlsx")

# Appending new values (by row) using the append() method

sheet = wb['Sheet 2']

sheet.append(['Test', 'value', '1'])

sheet.append(['Test', 'value', '2'])

wb.save("OpenPyXL practice workbook.xlsx")

# inserting and deleting rows

sheet.insert_rows(1)
sheet.delete_rows(2)

# inserting and deleting columns

sheet.insert_cols(1)
sheet.delete_cols(2)

wb.save("OpenPyXL practice workbook.xlsx")

data_sheet = wb['Data']

selected_range = data_sheet["A1":"G10"]

for cell in selected_range:
    for cell_no_tuple in cell:
        print(cell_no_tuple.value)