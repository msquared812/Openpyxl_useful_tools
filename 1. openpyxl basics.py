import openpyxl

wb = openpyxl.load_workbook("/Users/maxmaton/Library/Mobile Documents/com~apple~CloudDocs/Documents/Skills & Hobbies/Programming/Python/Open PyXL/OpenPyXL practice workbook.xlsx")

# Getting sheet names from workbook

print(wb.sheetnames)

# Selecting a worksheet and assigning it to a variable

sheet = wb['Sheet1']

# Selecting a cell and accessing its attributes:

print(sheet['A1'].value)
print(sheet['A1'].row)
print(sheet['A1'].column)

# selecting a cell using row and column number, rather than cell reference:

print(sheet.cell(1, 1).value)

# converting from column number to column letter:

print(openpyxl.utils.get_column_letter(sheet['A1'].column))

column_letter = openpyxl.utils.get_column_letter(sheet['A1'].column)

# and back again:

print(openpyxl.utils.column_index_from_string(column_letter))

# Looping over a range of cells

start_row = 1
end_row = 4

start_column = 1
end_column = 5

# alternative column selection using letter:

start_column_letter = openpyxl.utils.column_index_from_string('A')
end_column_letter = openpyxl.utils.column_index_from_string('E')

# the code will select first row and loop through the diff columns in that row, then move onto the next row and do the same:

for row in range(start_row, end_row + 1):
    for col in range(start_column, end_column + 1):
        print(sheet.cell(row, col).value)

# Looping over all the rows and columns in the worksheet:

for row in range(1, sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):
        print(sheet.cell(row, col).value)
